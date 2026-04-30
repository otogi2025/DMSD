"""食堂食数计算 + Excel 导出 (#7 / Q7)。

権威: system_features.md §7.7 + #38 (食事不要 from/to 明确)。
Q7 答え: 「Excel 表格、要包含的数据是哪些学生不需要餐食、什么期间。要可以一键导出 excel。」

ロジック:
1. 期間 [from, to] 内の各日付 D について、
2. status='approved' な applications で
   - meals_skip_from <= D の朝食時刻 <= meals_skip_to → 朝食 skip
   - 同様に 昼食・夕食 を判定
3. 出力 = 日別集計 + 学生別詳細 の 2 sheet

食事時間 (固定):
- 朝食 = 07:00 JST
- 昼食 = 12:00 JST
- 夕食 = 18:00 JST

⚠️ 2026-04-30 暫定値: 食堂運営の実時刻は老師に確認後、本ファイルの定数で調整。
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta, timezone
from typing import Iterable

from sqlalchemy import and_, select
from sqlalchemy.orm import Session

from .. import models

# 食堂 食事時刻 (JST 基準, 暫定)
JST = timezone(timedelta(hours=9))
BREAKFAST_AT = time(7, 0)
LUNCH_AT = time(12, 0)
DINNER_AT = time(18, 0)


@dataclass
class StudentMealDetail:
    student_no: str
    student_name: str
    dorm_unit: int
    room_no: str
    target_date: date
    breakfast_skip: bool
    lunch_skip: bool
    dinner_skip: bool


@dataclass
class DailyAggregate:
    target_date: date
    breakfast_skip: int = 0
    lunch_skip: int = 0
    dinner_skip: int = 0


@dataclass
class MealsCalcResult:
    range_from: date
    range_to: date
    daily: list[DailyAggregate] = field(default_factory=list)
    details: list[StudentMealDetail] = field(default_factory=list)

    @property
    def total(self) -> dict[str, int]:
        return {
            "breakfast_skip": sum(d.breakfast_skip for d in self.daily),
            "lunch_skip": sum(d.lunch_skip for d in self.daily),
            "dinner_skip": sum(d.dinner_skip for d in self.daily),
        }


# ---------------------------------------------------------------
# 計算
# ---------------------------------------------------------------
def _meal_dt(d: date, meal_at: time) -> datetime:
    return datetime.combine(d, meal_at, tzinfo=JST)


def _date_range(d_from: date, d_to: date) -> Iterable[date]:
    cur = d_from
    while cur <= d_to:
        yield cur
        cur += timedelta(days=1)


def calc_meals(
    db: Session, *, range_from: date, range_to: date
) -> MealsCalcResult:
    """[range_from, range_to] (両端含む) の食事不要数を計算。"""
    if range_to < range_from:
        raise ValueError("range_to must be >= range_from")

    # 期間と重なる承认済みの届を取る (帰省/外泊/帰国 全部 + meals_skip_from/to が NOT NULL)
    period_start = _meal_dt(range_from, BREAKFAST_AT) - timedelta(days=1)
    period_end = _meal_dt(range_to, DINNER_AT) + timedelta(days=1)

    stmt = (
        select(models.Application, models.Student)
        .join(models.Student, models.Student.id == models.Application.student_id)
        .where(
            and_(
                models.Application.status == "approved",
                models.Application.meals_skip_from.is_not(None),
                models.Application.meals_skip_to.is_not(None),
                models.Application.meals_skip_from <= period_end,
                models.Application.meals_skip_to >= period_start,
            )
        )
    )
    rows = db.execute(stmt).all()

    daily_map: dict[date, DailyAggregate] = {
        d: DailyAggregate(target_date=d) for d in _date_range(range_from, range_to)
    }
    details: list[StudentMealDetail] = []

    for app, student in rows:
        skip_from: datetime = app.meals_skip_from
        skip_to: datetime = app.meals_skip_to
        # tz-naive 入って来た場合は JST と解釈
        if skip_from.tzinfo is None:
            skip_from = skip_from.replace(tzinfo=JST)
        if skip_to.tzinfo is None:
            skip_to = skip_to.replace(tzinfo=JST)

        for d in _date_range(range_from, range_to):
            b_skip = _is_in_range(_meal_dt(d, BREAKFAST_AT), skip_from, skip_to)
            l_skip = _is_in_range(_meal_dt(d, LUNCH_AT), skip_from, skip_to)
            din_skip = _is_in_range(_meal_dt(d, DINNER_AT), skip_from, skip_to)
            if not (b_skip or l_skip or din_skip):
                continue
            agg = daily_map[d]
            agg.breakfast_skip += int(b_skip)
            agg.lunch_skip += int(l_skip)
            agg.dinner_skip += int(din_skip)
            details.append(
                StudentMealDetail(
                    student_no=student.student_no,
                    student_name=student.name,
                    dorm_unit=student.dorm_unit,
                    room_no=student.room_no,
                    target_date=d,
                    breakfast_skip=b_skip,
                    lunch_skip=l_skip,
                    dinner_skip=din_skip,
                )
            )

    return MealsCalcResult(
        range_from=range_from,
        range_to=range_to,
        daily=sorted(daily_map.values(), key=lambda x: x.target_date),
        details=sorted(
            details, key=lambda x: (x.target_date, x.student_no)
        ),
    )


def _is_in_range(d: datetime, lo: datetime, hi: datetime) -> bool:
    return lo <= d <= hi


# ---------------------------------------------------------------
# Excel 导出
# ---------------------------------------------------------------
def export_excel(result: MealsCalcResult) -> bytes:
    """openpyxl で .xlsx バイナリを返す。

    Sheet 1: 日別集計 (日付 / 朝 / 昼 / 夕 / 計)
    Sheet 2: 学生別詳細 (学号 / 名前 / 寮 / 部屋 / 日付 / 朝 / 昼 / 夕)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()

    # ---- Sheet 1: 日別集計 ----
    ws1 = wb.active
    ws1.title = "日別集計"
    ws1.append(
        ["日付", "曜日", "朝食 不要数", "昼食 不要数", "夕食 不要数", "合計"]
    )
    for c in ws1[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4472C4")
        c.alignment = Alignment(horizontal="center")

    weekday_jp = ["月", "火", "水", "木", "金", "土", "日"]
    for d in result.daily:
        total = d.breakfast_skip + d.lunch_skip + d.dinner_skip
        ws1.append(
            [
                d.target_date.isoformat(),
                weekday_jp[d.target_date.weekday()],
                d.breakfast_skip,
                d.lunch_skip,
                d.dinner_skip,
                total,
            ]
        )
    # 合計行
    totals = result.total
    grand_total = totals["breakfast_skip"] + totals["lunch_skip"] + totals["dinner_skip"]
    ws1.append(
        [
            "合計",
            "",
            totals["breakfast_skip"],
            totals["lunch_skip"],
            totals["dinner_skip"],
            grand_total,
        ]
    )
    last_row = ws1.max_row
    for c in ws1[last_row]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9E1F2")

    # 列幅
    widths = {"A": 12, "B": 6, "C": 14, "D": 14, "E": 14, "F": 10}
    for col, w in widths.items():
        ws1.column_dimensions[col].width = w

    # ---- Sheet 2: 学生別詳細 ----
    ws2 = wb.create_sheet(title="学生別詳細")
    ws2.append(
        ["日付", "学号", "氏名", "寮", "部屋", "朝食", "昼食", "夕食"]
    )
    for c in ws2[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4472C4")
        c.alignment = Alignment(horizontal="center")

    for det in result.details:
        ws2.append(
            [
                det.target_date.isoformat(),
                det.student_no,
                det.student_name,
                _dorm_label(det.dorm_unit),
                det.room_no,
                "○" if det.breakfast_skip else "",
                "○" if det.lunch_skip else "",
                "○" if det.dinner_skip else "",
            ]
        )

    widths2 = {
        "A": 12, "B": 10, "C": 18, "D": 12, "E": 8, "F": 8, "G": 8, "H": 8,
    }
    for col, w in widths2.items():
        ws2.column_dimensions[col].width = w
    for col in ("F", "G", "H"):
        for row in range(2, ws2.max_row + 1):
            ws2[f"{col}{row}"].alignment = Alignment(horizontal="center")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _dorm_label(dorm_unit: int) -> str:
    return {1: "1 寮", 2: "2 寮", 4: "4 寮"}.get(dorm_unit, f"{dorm_unit} 寮")


def export_filename(range_from: date, range_to: date) -> str:
    return f"meals_{range_from.isoformat()}_{range_to.isoformat()}.xlsx"
