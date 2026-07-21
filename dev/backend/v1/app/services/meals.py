"""食堂食数计算 + Excel 导出 (#7 / Q7)。

权威: system_features.md §7.7。
Q7 答え: 「Excel 表格、要包含的数据是哪些学生不需要餐食、什么期间。要可以一键导出 excel。」

逻辑:
1. 取 status='approved' 且 meals_skip 非空的 applications。
2. meals_skip = [{date, meal}, ...] 格式 — 1 条 = 1 餐不需要。
3. 把落在指定期间 [range_from, range_to] 内的条目按日汇总。
4. 输出 = 日别汇总 + 学生别详细 共 2 个 sheet。

⚠️ 2026-05-02 格式变更: 旧 meals_skip_from/to datetime range → 新 [{date, meal}] 列表。
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date, timedelta, timezone
from typing import Iterable

from sqlalchemy import and_, select
from sqlalchemy.orm import Session

from .. import models
from ..deps import demo_scope_for_teacher

JST = timezone(timedelta(hours=9))


@dataclass
class StudentMealDetail:
    student_no: str
    student_name: str
    dorm_unit: int
    room_no: str
    target_date: date
    breakfast_skip: bool
    lunch_skip: bool
    dinner_skip: bool


@dataclass
class DailyAggregate:
    target_date: date
    breakfast_skip: int = 0
    lunch_skip: int = 0
    dinner_skip: int = 0


@dataclass
class MealsCalcResult:
    range_from: date
    range_to: date
    daily: list[DailyAggregate] = field(default_factory=list)
    details: list[StudentMealDetail] = field(default_factory=list)

    @property
    def total(self) -> dict[str, int]:
        return {
            "breakfast_skip": sum(d.breakfast_skip for d in self.daily),
            "lunch_skip": sum(d.lunch_skip for d in self.daily),
            "dinner_skip": sum(d.dinner_skip for d in self.daily),
        }


# ---------------------------------------------------------------
# 计算
# ---------------------------------------------------------------
def _date_range(d_from: date, d_to: date) -> Iterable[date]:
    cur = d_from
    while cur <= d_to:
        yield cur
        cur += timedelta(days=1)


def calc_meals(
    db: Session, *, teacher: models.Teacher, range_from: date, range_to: date
) -> MealsCalcResult:
    """计算 [range_from, range_to]（含两端）期间内的不需餐食数。

    汇总 meals_skip = [{date: '2026-05-01', meal: '朝食'}, ...] 格式。

    teacher 的演示隔离: 用 demo_scope_for_teacher(teacher)
    让真老师只看真实学生 / 演示老师只看演示学生（防演示账号看到真实学生泄漏）。
    """
    if range_to < range_from:
        raise ValueError("range_to must be >= range_from")

    stmt = (
        select(models.Application, models.Student)
        .join(models.Student, models.Student.id == models.Application.student_id)
        .where(
            and_(
                models.Application.status == "approved",
                models.Application.meals_skip.is_not(None),
                # F-中-09：在 SQL 层按日期范围预筛，缩小候选集 —— meals_skip 条目只可能落在
                # 该申请的 [leave_date, return_date] 出寮窗口内，与查询区间无重叠的申请直接排除，
                # 不再全量 load approved 申请到内存逐条筛。重叠条件：窗口起点 <= 区间终点 且
                # 窗口终点 >= 区间起点。
                models.Application.leave_date <= range_to,
                models.Application.return_date >= range_from,
                demo_scope_for_teacher(teacher),
            )
        )
    )
    rows = db.execute(stmt).all()

    daily_map: dict[date, DailyAggregate] = {
        d: DailyAggregate(target_date=d) for d in _date_range(range_from, range_to)
    }
    # 用 (student_no, date) → StudentMealDetail 去重，避免同一人同一天重复计
    detail_map: dict[tuple, StudentMealDetail] = {}

    for app, student in rows:
        for entry in app.meals_skip or []:
            # F-中-08：entry 非 dict（脏数据 / 旧形式）时 entry['date'] 会抛 TypeError，
            # 原 except 只 catch (KeyError, ValueError) 漏掉它 → 整个导出 500。
            # 先 isinstance 判断跳过非 dict 条目，再处理键缺失 / 日期格式错。
            if not isinstance(entry, dict):
                continue
            try:
                entry_date = date.fromisoformat(entry["date"])
                meal = entry["meal"]
            except (KeyError, ValueError, TypeError):
                continue
            if entry_date not in daily_map:
                continue

            agg = daily_map[entry_date]
            key = (student.student_no, entry_date)
            if key not in detail_map:
                detail_map[key] = StudentMealDetail(
                    student_no=student.student_no,
                    student_name=student.name,
                    dorm_unit=student.dorm_unit,
                    room_no=student.room_no,
                    target_date=entry_date,
                    breakfast_skip=False,
                    lunch_skip=False,
                    dinner_skip=False,
                )
            det = detail_map[key]
            # 日别合计与明细同口径：同一 (学号, 日期, 餐次) 只计一次。
            # 明细靠 detail_map 按 (学号, 日期) 去重后再设 bool；合计也须在 bool
            # 从 False→True 时才 +1，否则跨申请/重复条目会二重计上。
            if meal == "朝食":
                if not det.breakfast_skip:
                    agg.breakfast_skip += 1
                det.breakfast_skip = True
            elif meal == "昼食":
                if not det.lunch_skip:
                    agg.lunch_skip += 1
                det.lunch_skip = True
            elif meal == "夕食":
                if not det.dinner_skip:
                    agg.dinner_skip += 1
                det.dinner_skip = True

    return MealsCalcResult(
        range_from=range_from,
        range_to=range_to,
        daily=sorted(daily_map.values(), key=lambda x: x.target_date),
        details=sorted(
            detail_map.values(), key=lambda x: (x.target_date, x.student_no)
        ),
    )


# ---------------------------------------------------------------
# Excel 导出
# ---------------------------------------------------------------
def export_excel(result: MealsCalcResult) -> bytes:
    """openpyxl で .xlsx バイナリを返す。

    Sheet 1: 日別集計 (日付 / 朝 / 昼 / 夕 / 計)
    Sheet 2: 学生別詳細 (学号 / 名前 / 寮 / 部屋 / 日付 / 朝 / 昼 / 夕)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()

    # ---- Sheet 1: 日別集計 ----
    ws1 = wb.active
    ws1.title = "日別集計"
    ws1.append(["日付", "曜日", "朝食 不要数", "昼食 不要数", "夕食 不要数", "合計"])
    for c in ws1[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4472C4")
        c.alignment = Alignment(horizontal="center")

    weekday_jp = ["月", "火", "水", "木", "金", "土", "日"]
    for d in result.daily:
        total = d.breakfast_skip + d.lunch_skip + d.dinner_skip
        ws1.append(
            [
                d.target_date.isoformat(),
                weekday_jp[d.target_date.weekday()],
                d.breakfast_skip,
                d.lunch_skip,
                d.dinner_skip,
                total,
            ]
        )
    # 合計行
    totals = result.total
    grand_total = (
        totals["breakfast_skip"] + totals["lunch_skip"] + totals["dinner_skip"]
    )
    ws1.append(
        [
            "合計",
            "",
            totals["breakfast_skip"],
            totals["lunch_skip"],
            totals["dinner_skip"],
            grand_total,
        ]
    )
    last_row = ws1.max_row
    for c in ws1[last_row]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9E1F2")

    # 列幅
    widths = {"A": 12, "B": 6, "C": 14, "D": 14, "E": 14, "F": 10}
    for col, w in widths.items():
        ws1.column_dimensions[col].width = w

    # ---- Sheet 2: 学生別詳細 ----
    ws2 = wb.create_sheet(title="学生別詳細")
    ws2.append(["日付", "学号", "氏名", "寮", "部屋", "朝食", "昼食", "夕食"])
    for c in ws2[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4472C4")
        c.alignment = Alignment(horizontal="center")

    for det in result.details:
        ws2.append(
            [
                det.target_date.isoformat(),
                det.student_no,
                det.student_name,
                _dorm_label(det.dorm_unit),
                det.room_no,
                "○" if det.breakfast_skip else "",
                "○" if det.lunch_skip else "",
                "○" if det.dinner_skip else "",
            ]
        )

    widths2 = {
        "A": 12,
        "B": 10,
        "C": 18,
        "D": 12,
        "E": 8,
        "F": 8,
        "G": 8,
        "H": 8,
    }
    for col, w in widths2.items():
        ws2.column_dimensions[col].width = w
    for col in ("F", "G", "H"):
        for row in range(2, ws2.max_row + 1):
            ws2[f"{col}{row}"].alignment = Alignment(horizontal="center")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _dorm_label(dorm_unit: int) -> str:
    return {1: "1 寮", 2: "2 寮", 4: "4 寮"}.get(dorm_unit, f"{dorm_unit} 寮")


def export_filename(range_from: date, range_to: date) -> str:
    return f"meals_{range_from.isoformat()}_{range_to.isoformat()}.xlsx"
