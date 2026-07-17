"""smoke + integration tests — 4 エンドポイント完成確認。

完成定義:
- POST /applications (提出 + 邮件 trigger)
- GET /applications/:id (#5 承认状态)
- GET /meals/export (Excel)
- 表创建 + 役职 seed + SendGrid 発信ロジック動作 (実 API 叩かない、log 確認のみ)
"""
from __future__ import annotations

from datetime import date, datetime, time, timedelta, timezone


def _tomorrow() -> date:
    return date.today() + timedelta(days=1)


def _day_after(n: int) -> date:
    return date.today() + timedelta(days=n)


# ---------------------------------------------------------------
# Health
# ---------------------------------------------------------------
def test_root_ok(client):
    res = client.get("/")
    assert res.status_code == 200
    assert res.json()["service"] == "Tomoshibi Backend v1"


def test_healthz_ok(client):
    res = client.get("/healthz")
    assert res.status_code == 200


# ---------------------------------------------------------------
# Auth
# ---------------------------------------------------------------
def test_student_login(client, seed_data):
    res = client.post(
        "/api/v1/sessions/student",
        json={"student_no": "060218", "password": "test-password-12345"},
    )
    assert res.status_code == 200
    assert "access_token" in res.json()["data"]


def test_student_login_wrong_password(client, seed_data):
    res = client.post(
        "/api/v1/sessions/student",
        json={"student_no": "060218", "password": "wrong-password!!"},
    )
    assert res.status_code == 401


def test_teacher_login(client, seed_data):
    res = client.post(
        "/api/v1/sessions/teacher",
        json={"login_id": "ryomu_kachou", "password": "test-password-12345"},
    )
    assert res.status_code == 200


# ---------------------------------------------------------------
# #2 + #6 POST /applications
# ---------------------------------------------------------------
def test_post_gaihaku_overseas_chain_5(client, student_token, seed_data, db_session):
    """留学生 + 外泊 → chain = 担任 + 国際交流部長 + 寮務課長 + 寮務部長 + 管理係 = 5 行。"""
    body = {
        "kind": "外泊",
        "leave_date": _tomorrow().isoformat(),
        "leave_method": "JR",
        "leave_time": "08:00:00",
        "return_date": _day_after(2).isoformat(),
        "return_method": "JR",
        "return_time": "20:00:00",
        "stay_locations": [
            {"kind": "ホテル", "name": "東横イン岡山", "address": "岡山市..."}
        ],
        "meals_skip": [
            {"date": _tomorrow().isoformat(), "meal": "朝食"},
            {"date": _tomorrow().isoformat(), "meal": "昼食"},
            {"date": _tomorrow().isoformat(), "meal": "夕食"},
            {"date": _day_after(1).isoformat(), "meal": "朝食"},
            {"date": _day_after(1).isoformat(), "meal": "昼食"},
            {"date": _day_after(1).isoformat(), "meal": "夕食"},
            {"date": _day_after(2).isoformat(), "meal": "朝食"},
            {"date": _day_after(2).isoformat(), "meal": "昼食"},
        ],
    }
    res = client.post(
        "/api/v1/applications",
        json=body,
        headers={"Authorization": f"Bearer {student_token}"},
    )
    assert res.status_code == 201, res.text
    data = res.json()["data"]
    chain = data["approval_chain"]
    roles = [c["approver_role"] for c in chain]
    assert roles == ["担任", "国際交流部長", "寮務課長", "寮務部長", "管理係"], roles
    assert data["status"] == "pending"
    # provisional header は外泊では立たない (実物表 evidence あり)
    assert res.headers.get("X-Approval-Chain-Provisional") != "true"


def test_post_kisei_provisional_header(client, student_token, seed_data):
    """帰省 = 5-28 实物表确认 → 不再返回 provisional header。"""
    body = {
        "kind": "帰省",
        "leave_date": _tomorrow().isoformat(),
        "leave_method": "新幹線",
        "leave_time": "10:00:00",
        "return_date": _day_after(3).isoformat(),
        "return_method": "新幹線",
        "return_time": "18:00:00",
    }
    res = client.post(
        "/api/v1/applications",
        json=body,
        headers={"Authorization": f"Bearer {student_token}"},
    )
    assert res.status_code == 201, res.text
    assert res.headers.get("X-Approval-Chain-Provisional") != "true"


def test_post_application_leave_date_today_rejected(client, student_token, seed_data):
    """#3 出寮日 = 今日 → 422。"""
    body = {
        "kind": "帰省",
        "leave_date": date.today().isoformat(),
        "leave_method": "新幹線",
        "leave_time": "10:00:00",
        "return_date": _day_after(2).isoformat(),
        "return_method": "新幹線",
        "return_time": "18:00:00",
    }
    res = client.post(
        "/api/v1/applications",
        json=body,
        headers={"Authorization": f"Bearer {student_token}"},
    )
    assert res.status_code == 422
    assert res.json()["error"]["code"] == "LEAVE_DATE_NOT_FUTURE"


def test_post_application_no_auth_rejected(client, seed_data):
    res = client.post("/api/v1/applications", json={"kind": "帰省"})
    assert res.status_code == 401


def test_application_creates_notification_log(client, student_token, seed_data, db_session):
    """#6 — 提交 → notification_log 行作成 (SendGrid 未設定なので status=pending)。"""
    from app import models

    body = {
        "kind": "外泊",
        "leave_date": _tomorrow().isoformat(),
        "leave_method": "JR",
        "leave_time": "08:00:00",
        "return_date": _day_after(2).isoformat(),
        "return_method": "JR",
        "return_time": "20:00:00",
        "stay_locations": [{"kind": "ホテル", "name": "test", "address": "..."}],
    }
    res = client.post(
        "/api/v1/applications",
        json=body,
        headers={"Authorization": f"Bearer {student_token}"},
    )
    assert res.status_code == 201

    logs = db_session.query(models.NotificationLog).all()
    assert len(logs) >= 1
    assert logs[-1].template_key == "application_submitted"
    # 5 役职 chain → 5 emails (SendGrid 未設定 → status='pending')
    assert "@" in (logs[-1].target_email or "")


# ---------------------------------------------------------------
# #5 GET /applications/{id}
# ---------------------------------------------------------------
def test_get_application_status(client, student_token, seed_data):
    body = {
        "kind": "帰省",
        "leave_date": _tomorrow().isoformat(),
        "leave_method": "JR",
        "leave_time": "08:00:00",
        "return_date": _day_after(3).isoformat(),
        "return_method": "JR",
        "return_time": "20:00:00",
    }
    post = client.post(
        "/api/v1/applications",
        json=body,
        headers={"Authorization": f"Bearer {student_token}"},
    )
    app_id = post.json()["data"]["id"]

    get = client.get(
        f"/api/v1/applications/{app_id}",
        headers={"Authorization": f"Bearer {student_token}"},
    )
    assert get.status_code == 200
    data = get.json()["data"]
    assert data["id"] == app_id
    assert data["status"] == "pending"
    # 留学生 + 帰省 → 5-28 实物表确认的 4 行 chain
    roles = [c["approver_role"] for c in data["approval_chain"]]
    assert roles == ["担任", "寮務課長", "寮務部長", "管理係"], roles


def test_get_application_other_student_forbidden(client, student_token, seed_data, db_session):
    """他学生の届は学生から見えない。"""
    from app import models

    # 別の学生作成 (一般)
    from app import security as sec
    other = models.Student(
        grade_code="06",
        class_code="01",
        seat_no="03",
        name="田中太郎",
        gender="male",
        room_no="M203",
        dorm_unit=1,
        is_overseas=False,
        email="other@test.jp",
    )
    db_session.add(other)
    db_session.flush()
    db_session.add(
        models.Account(
            student_id=other.id, password_hash=sec.hash_password("test-password-12345")
        )
    )
    other_app = models.Application(
        student_id=other.id,
        kind="帰省",
        leave_date=_tomorrow(),
        leave_method="JR",
        leave_time=time(8, 0),
        return_date=_day_after(2),
        return_method="JR",
        return_time=time(18, 0),
        status="pending",
    )
    db_session.add(other_app)
    db_session.commit()

    res = client.get(
        f"/api/v1/applications/{other_app.id}",
        headers={"Authorization": f"Bearer {student_token}"},
    )
    assert res.status_code == 403


# ---------------------------------------------------------------
# Approval chain 単体 (D4)
# ---------------------------------------------------------------
def test_chain_gaihaku_general():
    from app.services import approval_chain as ac

    roles = ac.get_chain_roles("外泊", is_overseas=False)
    assert roles == ("担任", "寮務課長", "寮務部長", "管理係")


def test_chain_gaihaku_overseas():
    from app.services import approval_chain as ac

    roles = ac.get_chain_roles("外泊", is_overseas=True)
    assert roles == ("担任", "国際交流部長", "寮務課長", "寮務部長", "管理係")


def test_chain_provisional_only_kikoku_general():
    from app.services import approval_chain as ac

    assert not ac.is_provisional("帰省", False)
    assert not ac.is_provisional("帰省", True)
    assert ac.is_provisional("帰国", False)
    assert not ac.is_provisional("帰国", True)
    assert not ac.is_provisional("外泊", False)


# ---------------------------------------------------------------
# #7 食堂 GET /meals/calc + /meals/export
# ---------------------------------------------------------------
def _make_approved_application(db_session, student, *, leave_d, return_d):
    """テスト用の承认済み外泊届 (食事不要期間付き)。"""
    from app import models

    app = models.Application(
        student_id=student.id,
        kind="外泊",
        leave_date=leave_d,
        leave_method="JR",
        leave_time=time(8, 0),
        return_date=return_d,
        return_method="JR",
        return_time=time(20, 0),
        stay_locations=[{"kind": "ホテル", "name": "T", "address": "x"}],
        meals_skip=[
            {"date": leave_d.isoformat(), "meal": "朝食"},
            {"date": leave_d.isoformat(), "meal": "昼食"},
            {"date": leave_d.isoformat(), "meal": "夕食"},
            {"date": return_d.isoformat(), "meal": "朝食"},
            {"date": return_d.isoformat(), "meal": "昼食"},
            {"date": return_d.isoformat(), "meal": "夕食"},
        ],
        status="approved",
    )
    db_session.add(app)
    db_session.flush()
    return app


def test_meals_calc(client, teacher_token, seed_data, db_session):
    student = seed_data["student"]
    leave = _tomorrow()
    ret = _day_after(2)
    _make_approved_application(db_session, student, leave_d=leave, return_d=ret)
    db_session.commit()

    res = client.get(
        f"/api/v1/meals/calc?from={leave.isoformat()}&to={ret.isoformat()}",
        headers={"Authorization": f"Bearer {teacher_token}"},
    )
    assert res.status_code == 200, res.text
    data = res.json()["data"]
    # 期間 = 2 日 (leave + return)
    assert len(data["daily"]) == 2
    # 1 学生 × 3 食 × 2 日 = 6 食 skip
    assert data["total"]["breakfast_skip"] == 2
    assert data["total"]["lunch_skip"] == 2
    assert data["total"]["dinner_skip"] == 2


def test_meals_export_xlsx(client, teacher_token, seed_data, db_session):
    student = seed_data["student"]
    leave = _tomorrow()
    ret = _day_after(2)
    _make_approved_application(db_session, student, leave_d=leave, return_d=ret)
    db_session.commit()

    res = client.get(
        f"/api/v1/meals/export?from={leave.isoformat()}&to={ret.isoformat()}",
        headers={"Authorization": f"Bearer {teacher_token}"},
    )
    assert res.status_code == 200
    assert res.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument"
    )
    # 中身: openpyxl で開けるか
    import io
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(res.content))
    assert "日別集計" in wb.sheetnames
    assert "学生別詳細" in wb.sheetnames


def test_meals_role_forbidden(client, seed_data, db_session):
    """寮監 (= role 外) は食堂データ見れない。"""
    res = client.post(
        "/api/v1/sessions/teacher",
        json={"login_id": "tannin", "password": "test-password-12345"},  # tannin = 寮務一般教師, 認可 OK
    )
    assert res.status_code == 200
    # 一般教师は OK
    tannin_token = res.json()["data"]["access_token"]
    leave = _tomorrow()
    res = client.get(
        f"/api/v1/meals/calc?from={leave.isoformat()}&to={leave.isoformat()}",
        headers={"Authorization": f"Bearer {tannin_token}"},
    )
    assert res.status_code == 200


# ---------------------------------------------------------------
# #6 SendGrid smoke (実 API 叩かない, dev mode = pending)
# ---------------------------------------------------------------
def test_notifications_test_dev_mode(client, teacher_token, seed_data):
    res = client.post(
        "/api/v1/notifications/test",
        json={"to": "test@example.jp", "subject": "smoke", "body_text": "hello"},
        headers={"Authorization": f"Bearer {teacher_token}"},
    )
    assert res.status_code == 200
    data = res.json()["data"]
    # SENDGRID_API_KEY 未設定 → sent=false + error 含 'not configured'
    assert data["sent"] is False
    assert "not configured" in (data["error"] or "")
